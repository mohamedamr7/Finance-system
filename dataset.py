import ollama
import os
import pandas as pd
import sys
import re
from datetime import datetime

# Import PDF and Report libraries at the top level to avoid NameErrors
try:
    from langchain_community.document_loaders import PyPDFLoader
    from langchain_text_splitters import RecursiveCharacterTextSplitter
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
except ImportError:
    print("\n[!] Critical: Missing libraries detected.")
    print("Please run: pip install langchain-community pypdf pandas openpyxl reportlab")
    sys.exit(1)

# --- CONFIGURATION ---
MODEL_NAME = "llama3" 
PDF_PATH = "imaginary_financial_dataset.pdf" 
SYSTEM_NAME = "financialAI Analysis"

def load_financial_data(file_path):
    """Loads the PDF silently."""
    if not os.path.exists(file_path):
        return []
    try:
        loader = PyPDFLoader(file_path)
        pages = loader.load()
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=800, chunk_overlap=50)
        return text_splitter.split_documents(pages)
    except Exception:
        return []

def export_to_excel(content, base_filename="Financial_Report"):
    """
    Creates a highly professional, regulated Excel sheet.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_filename}_{timestamp}.xlsx"
    
    try:
        lines = content.strip().split('\n')
        data_rows = []
        
        # Parse content into structured rows
        for line in lines:
            line = line.strip().replace("**", "")
            if not line or "---" in line: continue
            
            if "|" in line:
                cells = [c.strip() for c in line.split("|") if c.strip()]
                data_rows.append(cells)
            elif ":" in line:
                cells = [c.strip() for c in line.split(":", 1)]
                data_rows.append(cells)
            else:
                data_rows.append(["Analysis/Note", line])

        if not data_rows:
            data_rows = [["Summary", content]]

        df = pd.DataFrame(data_rows)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name='Analysis')
            
            worksheet = writer.sheets['Analysis']
            
            # Styling constants
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            row_fill_alt = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border_style = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # Apply styles
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row), start=1):
                for cell in row:
                    cell.border = border_style
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    if row_idx == 1: # Header
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif row_idx % 2 == 0: # Alternating rows
                        cell.fill = row_fill_alt

            # Auto-adjust column width
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                worksheet.column_dimensions[column].width = min(max_length + 5, 70)

        print(f"\n[System]: Excel file ready at: {os.path.abspath(filename)}")
    except PermissionError:
        print(f"\n[!] Error: Permission denied. Please close '{filename}' if it is open in Excel.")
    except Exception as e:
        print(f"\n[!] Excel Export Error: {e}")

def export_to_pdf(content, base_filename="Financial_Report"):
    """Creates a professional PDF report."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_filename}_{timestamp}.pdf"
    
    try:
        doc = SimpleDocTemplate(filename, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        story.append(Paragraph(f"{SYSTEM_NAME} - Analysis Report", styles['Title']))
        story.append(Spacer(1, 12))

        lines = content.strip().split('\n')
        for line in lines:
            clean = line.replace("**", "").strip()
            if clean:
                story.append(Paragraph(clean, styles['Normal']))
                story.append(Spacer(1, 6))

        doc.build(story)
        print(f"\n[System]: PDF ready at: {os.path.abspath(filename)}")
    except Exception as e:
        print(f"\n[!] PDF Export Error: {e}")

def stream_response(prompt, silent_files=False):
    """Streams response from the LLM."""
    full_response = ""
    if not silent_files:
        print("\n[AI]:", end=" ", flush=True)
    
    for chunk in ollama.generate(model=MODEL_NAME, prompt=prompt, stream=True):
        text = chunk['response']
        if not silent_files:
            print(text, end="", flush=True)
        full_response += text
    print()
    return full_response

def agent_router(user_query):
    prompt = f"Categorize query to ONE word: RPA, REPORT, PREDICTION. Query: {user_query}"
    response = ollama.generate(model=MODEL_NAME, prompt=prompt)
    return response['response'].strip().upper()

# --- MAIN INTERFACE ---
if __name__ == "__main__":
    # Clear screen for professional look
    os.system('cls' if os.name == 'nt' else 'clear')
    print("="*55)
    print(f"{SYSTEM_NAME.center(55)}")
    print("="*55)
    
    docs = load_financial_data(PDF_PATH)
    
    if docs:
        context_text = " ".join([d.page_content for d in docs[:4]]) 
        
        while True:
            try:
                user_input = input("\nHow can I help you? [exit to quit]: ")
                
                if user_input.lower() in ['exit', 'quit', 'q']:
                    print(f"\nShutting down {SYSTEM_NAME}...")
                    break
                
                if not user_input.strip(): continue
                
                is_file_req = any(x in user_input.lower() for x in ["excel", "pdf"])
                category = agent_router(user_input)
                
                # Use stream_response. If it's a file request, chatbot stays quiet.
                result = stream_response(f"Context: {context_text}\nTask: {user_input}", silent_files=is_file_req)
                
                if is_file_req:
                    print(f"[System]: Generating document based on your request...")
                    if "excel" in user_input.lower():
                        export_to_excel(result)
                    if "pdf" in user_input.lower():
                        export_to_pdf(result)
                
            except KeyboardInterrupt:
                print("\nSession interrupted.")
                break
    else:
        print("\n[!] Critical: System could not load financial data. Ensure your PDF is in the folder.")